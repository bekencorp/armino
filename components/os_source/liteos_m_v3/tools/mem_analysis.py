#!/usr/bin/env python
# -*- coding: utf-8 -*-

#
# Copyright (c) 2020 Huawei Device Co., Ltd.
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#

import re
import sys
import argparse
import os
import datetime

g_version = 2.5
g_excel_support = False
g_row_num = 0

try:
    from openpyxl import Workbook
except:
    print ("excel output is not support, need install openpyxl")
else:
    g_excel_support = True
    
def is_number(num):
    try:
        int(num, 16)
        return True
    except ValueError:
        pass
    return False

static_map = {}
def storage_static_data(offset, section, sizeHex, symbol, lib, obj):
    global g_row_num
    if (0 == g_row_num):
        static_map[g_row_num] = {'offsets' : "offsets", 'section' : "section",\
            'sizeH' : "size(Hex)", 'sizeD' : "size(Dec)",\
            'symbol' : "symbol", 'lib' : "lib", 'obj' : "object"}
        g_row_num = g_row_num + 1
    
    size = int(sizeHex, 16)
    static_map[g_row_num] = {'offsets': offset, 'section' : section,\
            'sizeH' : sizeHex, 'sizeD' : size,\
            'symbol' : symbol, 'lib' : lib, 'obj' : obj}
    g_row_num = g_row_num + 1
 
def store_static_excel():
    wb = Workbook()
    ws = wb.active
    for base_address, values in static_map.items():
        c = ws.cell(row = base_address + 1, column = 1)
        c.value = values.get('offsets')
        c = ws.cell(row = base_address + 1, column = 2)
        c.value = values.get('section')
        c = ws.cell(row = base_address + 1, column = 3)
        c.value = values.get('sizeH')
        c = ws.cell(row = base_address + 1, column = 4)
        c.value = values.get('sizeD')
        c = ws.cell(row = base_address + 1, column = 5)
        c.value = values.get('symbol')
        c = ws.cell(row = base_address + 1, column = 6)
        c.value = values.get('lib')
        c = ws.cell(row = base_address + 1, column = 7)
        c.value = values.get('obj')  
    wb.save('static_symbol-%s.xlsx' %datetime.datetime.now().strftime('%Y-%m-%d %H_%M_%S'))
    return

def format_store_static():
    global g_excel_support
    out_fd = open("./static_symbol_out.txt", "w+")
    target_list = []
    
    for base_address, values in static_map.items():
        target_list.append('{:<10s}'.format(values.get('offsets')))
        target_list.append('{:<30s}'.format(values.get('section')))
        target_list.append('{:<10}'.format(values.get('sizeH')))
        target_list.append('{:<10}'.format(values.get('sizeD')))
        target_list.append('{:<55s}{:<55s}{:<55s}'.format(values.get('symbol'), values.get('lib'), values.get('obj')))
        target_list.append("\r")
        
    out_fd.write("".join(target_list))
    if g_excel_support == True:
        store_static_excel()
    return
    
dync_map = {}
#dync map multiple LR, first dync_map node will record LR size 
def storage_dync_head(size):
    global g_row_num
    dync_map[0] = {'addr' : "Memnode", 'symbol' : "size", 'lib' : size}
    g_row_num = 1

#when the analysis of one line LR is finished, need record "\r\n" and the new line info
def storage_dync_node(node_addr, size):
    global g_row_num
    dync_map[g_row_num] = {'addr' : "\r\n", 'symbol' : node_addr, 'lib' : size}
    g_row_num = g_row_num + 1

def storage_dync_data(addr, symbol, lib):
    global g_row_num
    dync_map[g_row_num] = {'addr' : addr, 'symbol' : symbol, 'lib' : lib}
    g_row_num = g_row_num + 1
    
def store_dync_excel():
    wb = Workbook()
    ws = wb.active
    row = 1
    column = 0
    for base_address, values in dync_map.items():
        if (0 == base_address):
            c = ws.cell(row, 1)
            c.value = values.get('addr') 
            c = ws.cell(row, 2)
            c.value = values.get('symbol')
            for i in range(values.get('lib')):
                column = i * 3 + 2
                addr_name = 'LR' + str(i) + '_addr'
                c = ws.cell(row, column + 1)
                c.value = addr_name                
                symbol_name = 'LR' + str(i) + '_symbol'
                c = ws.cell(row, column + 2)
                c.value = symbol_name   
                object_name = 'LR' + str(i) + '_object'
                c = ws.cell(row, column + 3)
                c.value = object_name
        else:
            if("\r\n" == values.get('addr')):
                row = row + 1
                column = 1
                c = ws.cell(row, column)
                c.value = values.get('symbol')
                column = column + 1
                c = ws.cell(row, column)
                c.value = values.get('lib')
            else:
                c = ws.cell(row, column + 1)
                c.value = values.get('addr')
                c = ws.cell(row, column + 2)
                c.value = values.get('symbol')
                c = ws.cell(row, column + 3)
                c.value = values.get('lib') 
                column = column + 3
    wb.save('dync_mem-%s.xlsx' %datetime.datetime.now().strftime('%Y-%m-%d %H_%M_%S'))
    return

def format_store_dync():
    global g_excel_support
    out_fd = open("./dynamic_memory_out.txt", "w+")
    target_list = []
    
    for base_address, values in dync_map.items():
        if (0 == base_address):
            target_list.append('{:<15s}{:<15s}'.format("node", "size"))
            for i in range(values.get('lib')):
                addr_name = 'LR' + str(i) + '_addr'
                symbol_name = 'LR' + str(i) + '_symbol'
                object_name = 'LR' + str(i) + '_object'
                target_list.append('{:<15s}{:<55s}{:<55s}'.format(addr_name, symbol_name, object_name))
                
        else:
            if ("\r\n" == values.get('addr')):
                target_list.append("\r")
                target_list.append('{:<15s}{:<15}'.format(values.get('symbol'), values.get('lib')))
            else:
                target_list.append('{:<15s}'.format(values.get('addr')))
                target_list.append('{:<55s}'.format(values.get('symbol')))
                target_list.append('{:<55s}'.format(values.get('lib')))
        
    out_fd.write("".join(target_list))
    if g_excel_support == True:
        store_dync_excel()
    return

address_map = {}
def get_func_by_name(function_name):
    for base_address, values in address_map.items():
        if values.get('func_name') == function_name:
            return values.get('lib_name'), values.get('obj_name')
    return "none", "none" 

def get_func_by_address(address):
    if 0 == address:
        return "none", "none"

    for base_address, values in address_map.items():
        if (base_address <= address) and (base_address  + values.get('size') > address):
            return values.get('func_name'), values.get('obj_name')
    return "none", "none"
    
def line_start_check(line):
    if line.startswith(' .text.'):
        function_name = line.split(".text.")[1]
    elif line.startswith(' .rodata.'):
        function_name = line.split(".rodata.")[1]
    elif line.startswith(' .bss.'):
        function_name = line.split(".bss.")[1]
    elif line.startswith(' .sram.text.'):
        function_name = line.split(".sram.text.")[1]
    elif line.startswith(' .data.'):
        function_name = line.split(".data.")[1]
    else:
        return None
    return function_name
    
def parse_map(file_path):
    parse_begin = False
    function_name = ""
    address = None
    size = None
    obj_name = ""
    lib_name = ""
    pl = re.compile(r'[(](.*?)[)]', re.S)
    with open(file_path, "r") as f:
        for line in f.readlines():
            if not parse_begin:
                function_name = line_start_check(line)
                if function_name != None:
                    if len(function_name.split()) > 1:
                        strs = function_name.split()
                        function_name = strs[0]
                        function_name = function_name.strip()
                        address = int(strs[1], 16)
                        size = int(strs[2], 16)
                        obj_name = re.findall(pl, strs[3])
                        if len(obj_name) > 0:
                            obj_name = obj_name[0]
                        else:
                            obj_name = strs[3]
                        char_temp = strs[3].split('(')
                        if len(char_temp) > 1:
                            lib_name = char_temp[0].split('/')[-1]
                        else:
                            lib_name = "none"
                        address_map[address] = {'size' : size, 'lib_name' : lib_name, \
                            'obj_name' : obj_name, 'func_name' : function_name}
                    else:
                        parse_begin = True
                else:
                    continue
            else:
                parse_begin = False
                if len(line.strip().split()) != 3:
                    continue
                strs = line.strip().split()
                address = int(strs[0], 16)
                size = int(strs[1], 16)
                obj_name = re.findall(pl, strs[2])
                if len(obj_name) > 0:
                    obj_name = obj_name[0]
                else:
                    obj_name = strs[2]
                function_name = function_name.strip()
                char_temp = strs[2].split('(')
                if len(char_temp) > 1:
                    lib_name = char_temp[0].split('/')[-1]
                else:
                    lib_name = "none"
                address_map[address] = {'size' : size, 'lib_name' : lib_name, \
                    'obj_name' : obj_name, 'func_name' : function_name}

def parse_dump(dump_path):
    with open(dump_path, "r") as symbol_fd:
        symline = symbol_fd.readlines()
    
    for index, line in enumerate(symline):
        offset = line[:8]
        
        #According to the regular expression of objdump , section start at 17 char  
        symbol_list = line[17:].split()
        if len(symbol_list) < 3:
            continue
        
        sec = symbol_list[0]
        size = symbol_list[1]
        if (False == is_number(size)) or 0 == int(size, 16):
            continue

        lib, obj = get_func_by_name(symbol_list[-1])
        storage_static_data(offset, sec, size, symbol_list[-1], lib, obj)

def parse_log(log_path):
    with open(log_path, "r") as log_fd:
        logline = log_fd.readlines()
    get_valid_log = False
    for index, line in enumerate(logline):
        char_list = line.split()
        if len (char_list) < 3:
            continue
        if ("node" == char_list[0]) & ("LR[0]" == char_list[2]):
            storage_dync_head(len(char_list) - 2)
            get_valid_log = True
            continue
            
        if True == get_valid_log:
            mem_addr = char_list[0].split(':')
            if True == is_number(mem_addr[0]):
                for i in range(len(char_list)):
                    if i == 0:
                        size = int(char_list[1], 16)
                        storage_dync_node(mem_addr[0], size)
                    if i > 1:
                        symbol, lib = get_func_by_address(int(char_list[i], 16))
                        storage_dync_data(char_list[i], symbol, lib)
            else:
                get_valid_log = False
                continue

def main():
    print ("memory parses tool ver.%2f\r\n" %g_version)
    parser = argparse.ArgumentParser()
    parser.add_argument('--m', help = 'map path.')
    parser.add_argument('--l', help = 'dynamic mem log path.')    
    parser.add_argument('--d', help = 'objdump path.')
    args = parser.parse_args()
    
    print ("map path: %s\r\n" %args.m)
    if args.m == None :
        print ("arg error, input -h get the help list\r\n")
        return

    parse_map(args.m)

    print ("dump path: %s" %args.d)
    if args.d != None :
        parse_dump(args.d)
        format_store_static()
    else:
        print ("dump path unspecified, will not be static parser")
        print ("you can enter the objdump -t command under the linux shell to obtain dump information\r\n")
    
    print ("log path: %s" %args.l)
    if args.l != None :
        parse_log(args.l)
        format_store_dync()
    else:
        print ("log path unspecified, will not be dynamic parser\r\n")

    return

if __name__ == '__main__':
    sys.exit(main())